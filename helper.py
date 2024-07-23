import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

folder = '/Users/lucycai/Downloads/Orro_Bills'
def excel_2_df(src_excel, sheet_name):
    df = pd.read_excel(src_excel, sheet_name=sheet_name)
    return df

def highlight_excel(output_file_path):
    # Load the Excel workbook and select a worksheet
    wb = load_workbook(output_file_path)
    ws = wb.active
    ws.title = 'summary'

    # Define a fill pattern for highlighting
    highlight = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Apply highlight to each summary row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Assuming first row is header
        if row[4].value == 'Charge Back Journal':  # Check if it's a summary row
            for cell in row:
                cell.fill = highlight

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Add extra space for better visibility
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook
    wb.save(os.path.join(folder, 'highlighted.xlsx'))

def export_2_excel(output_file_path,df,*args):

    with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
        sheetname = 'comparison'
        df.to_excel(writer, sheet_name=sheetname, index=False)

        workbook = writer.book
        num_format = workbook.add_format({'num_format': '0'})
        worksheet = writer.sheets[sheetname]

        # Adjust width of columns based on header length
        for col_num, column in enumerate(df.columns):
            
            max_length = max(df[column].astype(str).map(len).max(), len(column))  # Max length of content or header
        
            if column == args[0]:
                worksheet.set_column(col_num,col_num,max_length + 2, num_format)
                continue
            # Set the column width based on header length with a bit of padding
            worksheet.set_column(col_num, col_num, max_length)  # Adding 2 for padding

